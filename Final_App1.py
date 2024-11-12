from flask import Flask, render_template, request, redirect, url_for
from flask_cors import CORS
from flask_socketio import SocketIO, emit
from configparser import ConfigParser
from datetime import datetime, timedelta
import csv
#from csv import delimiter, quotechar
import pandas as pd
import os
import shutil
import openpyxl
app = Flask(__name__)
socketio = SocketIO(app)
CORS(app)


@socketio.on('sendmessage')
def handle_print_message(message):
    print(message)
    socketio.emit('status message', message)

@socketio.on('connect')
def handle_connect():
    print('Client connected')
    emit('after connect',  'Lets dance')
    

Base_dir = os.path.dirname(os.path.abspath(__file__))
UPLOAD_FOLDER = os.path.join(Base_dir, 'uploads')
Input_folder = os.path.join(UPLOAD_FOLDER, 'Input')
Backup_folder = os.path.join(UPLOAD_FOLDER, 'Backup')
Archives_folder = os.path.join(UPLOAD_FOLDER, 'Archives')
Error_folder = os.path.join(UPLOAD_FOLDER, 'Error')
ALLOWED_EXTENSIONS = {'xls', 'xlsx'}
app.config['UPLOAD_FOLDER'] = Archives_folder

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS
import os
import pandas as pd
import csv
from datetime import datetime, timedelta

import os
import pandas as pd
from datetime import datetime, timedelta


import os
import pandas as pd
from datetime import datetime, timedelta

def export(filename):
    handle_print_message(f"Started processing file {filename}")
    current_file = pd.read_excel(filename)
    # defining the header row of the passed excel file
    header_row = current_file.columns
    first_row = 0
    current_row = first_row
    last_row = current_file.shape[0]
    allocation_number = read_allocation_number()
    if allocation_number >= 999999:
        allocation_number = 1
    else:
        allocation_number += 1
    write_ini_file_string("Allocation", "AllocationNumber", str(allocation_number))
    csvPath = os.path.join(Base_dir, 'extract')
    csvname = f"MA{allocation_number:07d}.dat"
    csvname = os.path.join(csvPath, csvname)
    with open(csvname, 'w',newline='') as csvFile:
        writer = csv.writer(csvFile, quotechar=None)
        #writer.writerow(get_csv_header())
        
        if first_row <= last_row:
            for current_row in range(first_row, last_row):
                print(f"Writing Replenishment Record from row {current_row+1}")
                row_data = get_csv_row(current_row,filename)
                writer.writerow(row_data)
        else:
            print("No data found in the file.")
        print(f"File {csvname} has been created.")
        csvFile.close()
    move_file_backup(filename,str(allocation_number))
def get_csv_header():
    return ["Class_Style_Color_Size", "AEO_DC_From_LOC", "AEO_DC_To_LOC", "Allocate_Date", "Ship_Date", 
            "Arrive_Date", "BULK_OH_Class_Style_Color_Size_DC", 
            "BULK_OH_Class_Style_Color_Size_Alloc", "Class_Style_Color_Size", 
            "", "QTY", "0", "0", "Alloc", "1", "N", "", "DEPLOYMENT", "0", "Manual Import", "0000000000", 
            "FloorSet", "Current_Time_Stamp"]

def get_csv_row(current_row,filename):
    current_file = pd.read_excel(filename)
    from_warehouse = format(current_file.iloc[current_row,0], "000000")
    from_warehouse = from_warehouse.zfill(6)
    to_warehouse = format(current_file.iloc[current_row,1], "000000")
    to_warehouse = to_warehouse.zfill(6)
    dept = current_file.iloc[current_row,2]
    class_val = format(current_file.iloc[current_row,3], "0000")
    class_val = class_val.zfill(4)
    style = format(current_file.iloc[current_row,4], "0000")
    style = style.zfill(4)
    color = format(current_file.iloc[current_row,5], "000")
    color = color.zfill(3)
    size = format(current_file.iloc[current_row,6], "0000")
    size = size.zfill(4)
    qty = current_file.iloc[current_row,7]
    alloc_num = format(int(read_allocation_number()), "0000000")
    alloc_num = alloc_num.zfill(7)
    rowData = [
        f'"{class_val}_{style}_{color}_{size}"',
        f'"AEO_DC_{from_warehouse}"',
        f'"AEO_ST_{to_warehouse}"',
        f'"{datetime.now().strftime("%Y%m%d")}"',
        f'"{(datetime.now() + timedelta(days=1)).strftime("%Y%m%d")}"',
        f'"{(datetime.now() + timedelta(days=4)).strftime("%Y%m%d")}"',
        f'"BULK_OH_{class_val}_{style}_{color}_{size}_DC{from_warehouse}"',
        f'"BULK_OH_{class_val}_{style}_{color}_{size}_DC{from_warehouse}_ALC{alloc_num}"',
        f'"{class_val}_{style}_{color}_{size}"',
        '""',
        f'"{qty}"',
        '"0"',
        '"0"',
        f'"{alloc_num}"',
        '"1"',
        '"N"',
        '""',
        '"DEPLOYMENT"',
        '"0"',
        '"Manual Import"',
        '"0000000000"',
        '"  "',
        f'"{datetime.now().strftime("%Y-%m-%d.%H.%M.%S")}"'
    ]

    return rowData

#function to rename the file and move file from Input folder to Error folder
def move_error_file(filename):
    #rename filename by adding current date & time
    current_date = datetime.now().strftime("%Y%m%d")
    current_time = datetime.now().strftime("%H%M%S")
    #new_filename should be filename without extension + current date + current time
    new_filename = filename.split('.')[0] + "_" + current_date + "_" + current_time + "." + filename.split('.')[1]
    current_file = os.path.join(Input_folder, filename)
    #move file to Error folder
    shutil.move(current_file, Error_folder)
    #rename file
    os.rename(os.path.join(Error_folder, filename), os.path.join(Error_folder, new_filename))
    print("File moved to Error folder")

def move_file_backup(filename,allocation_number):
    new_filename = filename.split('.')[0] + "_" + allocation_number + "." + filename.split('.')[1]
    print(f"new file name is {new_filename}")
    current_file = os.path.join(Input_folder, filename)
    print(f"current file is {current_file}")
    os.rename(current_file, new_filename)
    print(f"file renamed to {new_filename}")
    #move file to Error folder
    shutil.move(new_filename, Backup_folder)
    print(f"file moved from {new_filename} to {Backup_folder}")
    #rename file
    print("Exported File moved to Backup folder")

def split_excel(filename):
    for sheet in pd.read_excel(filename, sheet_name=None):
        df = pd.read_excel(filename, sheet_name=sheet)
        df.to_excel(os.path.join(Input_folder, sheet + '.xlsx'), index=False)
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def ini_file_name():
    return os.path.join(os.getcwd(), "Allocation.ini")

def read_allocation_number():
    ini_file = ini_file_name()

    # Check if the file exists
    if not os.path.exists(ini_file):
        print(f"INI file '{ini_file}' not found.")
        return None

    config = ConfigParser()
    config.read(ini_file)

    try:
        return config.getint("Allocation", "AllocationNumber")
    except Exception as e:
        print(f"Error reading AllocationNumber from INI file: {e}")
        return None

def write_ini_file_string(section, keyname, wstr):
    ini_file = ini_file_name()

    # Check if the file exists
    if not os.path.exists(ini_file):
        print(f"INI file '{ini_file}' not found.")
        return ""

    config = ConfigParser()
    config.read(ini_file)

    try:
        config.set(section, keyname, wstr)

        # Writing the changes to the INI file
        with open(ini_file, 'w') as config_file:
            config.write(config_file)

        return wstr
    except Exception as e:
        print(f"Error writing to INI file: {e}")
        return ""
    
def find_errors(filename):
    Field_File = os.path.join(Base_dir, 'Field_Details.xlsx')
    Lookup_Ref = openpyxl.load_workbook(Field_File)
    Lookup_Ref_sheet = Lookup_Ref['Lookups']
    current_file = openpyxl.load_workbook(filename)
    cws = current_file.active

    # defining the header row of the passed excel file
    header_row = cws[1]
    first_row = header_row[0].row + 1
    current_row = first_row
    cws_column = cws['C']
    last_row = max(cell.row for cell in cws_column if cell.value is not None)
    no_errors = True
    if last_row >= first_row:
        No_of_Columns = cws.max_column
        #print("number of columns in input file is ", No_of_Columns)
        column_names = Lookup_Ref_sheet['A2':'A9']
        number_of_columns = len(column_names)
        #print("number of columns required is ", number_of_columns)
        if No_of_Columns == number_of_columns:
            
            while current_row <= last_row: 

                for i in range(number_of_columns):
                    current_column = openpyxl.utils.get_column_letter(i + 1)
                    #print("current column is ",current_column)
                    column_name = cws[current_column + str(header_row[0].row)].value
                    #print("column name is ", column_name)
                    field_value = str(cws[current_column + str(current_row)].value)
                    #print("field value is ", field_value)
                    column_value = Lookup_Ref_sheet['B2':'B9'][i][0].value
                    
                    #print("column value is ", column_value)
                    field_required = Lookup_Ref_sheet['E2':'E9'][i][0].value

                    if field_value is not None:
                        field_type = Lookup_Ref_sheet['C2':'C9'][i][0].value
                        field_length = Lookup_Ref_sheet['D2':'D9'][i][0].value

                        #check data type with Lookup sheet
                        if (len(field_value) > field_length or \
                            (field_type.upper() == 'NUMBER' and not field_value.isnumeric()) or \
                            (field_type.upper() == 'DATE' and not isinstance(field_value, (str,int, float)))):
                            print(f"{current_column} {current_row} is not a valid data type. Data type includes length of fields.")
                            socketio.emit('print_message', f"{current_column} {current_row} is not a valid data type. Data type includes length of fields.")
                            no_errors = False
                        
                        # Pad value out with 0's
                        if field_type is not None and field_type.upper() == 'VARCHAR2' and len(field_value) < field_length:
                            field_value = field_value.rjust(field_length, '0')
                            cws[current_column + str(current_row)].value = field_value

                    else:
                        if field_required == 'Y':
                            print(f"{current_column}{current_row} is a required field. A value must be entered.")
                            socketio.emit('print_message', f"{current_column}{current_row} is a required field. A value must be entered.")
                            no_errors = False

                current_row += 1
            if no_errors:
                print(f"No errors found in the file.{filename}")
                socketio.emit('print_message', {'message': 'No errors found in the file.'})
                export(filename)

        else:
            print("Number of columns in the input file is not matching with the required columns")
            socketio.emit('print_message', "Number of columns in the input file is not matching with the required columns")
            move_error_file(filename)
    else:
        print("No data found in the file.")
        socketio.emit('print_message', {'message': 'No data found in the file.'})
        move_error_file(filename)

@app.route('/')
def index():
    return render_template('index.html')
@app.route('/file-validation')
def File_Validation():
    return render_template('File_Validation.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return redirect(request.url)

    file = request.files['file']

    if file.filename == '':
        return redirect(request.url)

    if file and allowed_file(file.filename):
        filename = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
        file.save(filename)
        split_excel(filename)
        for each_file in os.listdir(Input_folder):
            if each_file.endswith(".xlsx"):
                file_path = os.path.join(Input_folder, each_file)
                find_errors(file_path)


        #return redirect(url_for('File_Validation'))
                
   
        return redirect(url_for('index'))
        
    #return redirect(request.url)

if __name__ == '__main__':
    #app.run(debug=True, port=8080) c:\Users\rvarasala\Downloads\Allocation.ini
    app.run(host='0.0.0.0', port=8080)
    socketio.run(app)

#find_errors('Upload 1.xlsx')